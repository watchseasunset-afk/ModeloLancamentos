"""
Análise estatística das variáveis que influenciam Lançamentos Laterais — PL 2025/26

Métodos:
  1. Correlações de Pearson (cada variável vs lançamentos)
  2. Regressão Linear Múltipla (total lançamentos por jogo)
  3. Análise Over/Under por limiares
  4. Perfis de equipa por categoria tática

Corre: python analise_variaveis.py
Output: data/analise_lancamentos_PL_2025_26.xlsx
"""
import sys, subprocess, site
for pkg in ['pandas','openpyxl','scipy','scikit-learn','numpy']:
    try:
        __import__(pkg.replace('scikit-learn','sklearn'))
    except ImportError:
        subprocess.check_call([sys.executable,'-m','pip','install',pkg,'--break-system-packages'])
# Garantir site-packages do utilizador
for sp in site.getusersitepackages() if hasattr(site,'getusersitepackages') else []:
    if sp not in sys.path: sys.path.insert(0, sp)
import importlib
for mod in ['scipy','sklearn']:
    try: importlib.import_module(mod)
    except ImportError:
        for p in ['/sessions/charming-hopeful-davinci/.local/lib/python3.10/site-packages']:
            if p not in sys.path: sys.path.insert(0, p)

import numpy as np
import pandas as pd
from scipy import stats
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import r2_score
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from pathlib import Path

DATA_DIR   = Path(__file__).parent / 'data'
STATS_FILE = DATA_DIR / 'stats_completas_PL_2025_26.xlsx'
TEAMS_MAP  = Path(__file__).parent / 'teams_map.csv'
OUT_FILE   = DATA_DIR / 'analise_lancamentos_PL_2025_26.xlsx'

import re

def parse_stat(val):
    """Parse 'XX% (Y/Z)' → denominator Z; ou numérico simples."""
    if pd.isna(val):
        return np.nan
    s = str(val).strip()
    m = re.search(r'\((\d+)/(\d+)\)', s)
    if m:
        return float(m.group(2))   # total tentativas (denominador)
    try:
        return float(s.replace('%','').strip())
    except:
        return np.nan

print("A carregar dados...")
df    = pd.read_excel(STATS_FILE)
teams = pd.read_csv(TEAMS_MAP)
df    = df.merge(teams, on='mid', how='left')

# Converter tudo para numérico — colunas "X% (Y/Z)" extraem o denominador
for c in df.columns:
    if c not in ('mid','home','away'):
        df[c] = df[c].apply(parse_stat)

# ── Criar variáveis de análise por equipa/jogo ─────────────────────────────────
# Perspetiva de cada equipa em cada jogo (linha = 1 equipa num jogo)
rows = []
for _, r in df.iterrows():
    for side, opponent, team_col, opp_col in [('casa','fora','home','away'),('fora','casa','away','home')]:
        row = {
            'mid':      r['mid'],
            'equipa':   r[team_col],
            'adversario': r[opp_col],
            'lado':     side,
            # TARGET
            'lancamentos': r[f'Lançamentos_{side}'],
            # GRUPO 1: Estilo de jogo / construção
            'passes':          r[f'Passes_{side}'],
            'passes_longos':   r[f'Passes longos_{side}'],
            'passes_ult3':     r[f'Passes no último terço_{side}'],
            'passes_prof':     r[f'Passes em profundidade certos_{side}'],
            'cruzamentos':     r[f'Cruzamentos_{side}'],
            'posse':           r[f'Posse de bola_{side}'],
            # GRUPO 2: Pressão defensiva adversária (o que o adversário faz a nós)
            'faltas_adv':      r[f'Faltas_{opponent}'],
            'desarmes_adv':    r[f'Desarmes_{opponent}'],
            'intercecoes_adv': r[f'Interceções_{opponent}'],
            # GRUPO 3: Nossa defesa / reação
            'alivios':         r[f'Alívios_{side}'],
            'faltas':          r[f'Faltas_{side}'],
            'desarmes':        r[f'Desarmes_{side}'],
            'intercecoes':     r[f'Interceções_{side}'],
            # GRUPO 4: Intensidade / contexto
            'duelos':          r[f'Duelos ganhos_{side}'],
            'foras_jogo':      r[f'Foras de Jogo_{side}'],
            'livres':          r[f'Livres_{side}'],
            # GRUPO 5: Ataque / largura
            'toques_area':     r[f'Toques na área adversária_{side}'],
            'remates':         r[f'Total remates_{side}'],
        }
        rows.append(row)

ap = pd.DataFrame(rows).dropna(subset=['lancamentos'])
print(f"  {len(ap)} linhas (equipa × jogo)")

# Total lançamentos por jogo
df['total_lanc'] = df['Lançamentos_casa'] + df['Lançamentos_fora']
df['total_passes'] = df['Passes_casa'] + df['Passes_fora']
df['total_passes_longos'] = df['Passes longos_casa'] + df['Passes longos_fora']
df['total_posse_diff'] = (df['Posse de bola_casa'] - df['Posse de bola_fora']).abs()
df['total_cruzamentos'] = df['Cruzamentos_casa'] + df['Cruzamentos_fora']
df['total_faltas'] = df['Faltas_casa'] + df['Faltas_fora']
df['total_alivios'] = df['Alívios_casa'] + df['Alívios_fora']
df['total_desarmes'] = df['Desarmes_casa'] + df['Desarmes_fora']
df['total_intercecoes'] = df['Interceções_casa'] + df['Interceções_fora']
df['total_duelos'] = df['Duelos ganhos_casa'] + df['Duelos ganhos_fora']

print(f"\n  Média lançamentos/jogo: {df['total_lanc'].mean():.1f}")
print(f"  Mediana: {df['total_lanc'].median():.0f} | Std: {df['total_lanc'].std():.1f}")
print(f"  Range: {df['total_lanc'].min():.0f} – {df['total_lanc'].max():.0f}")

# ════════════════════════════════════════════════════════════════════════════════
# ANÁLISE 1: Correlações (perspetiva por equipa)
# ════════════════════════════════════════════════════════════════════════════════
print("\nA calcular correlações...")

variaveis = {
    # (nome_display, coluna, grupo, interpretação)
    'Passes (total)':             ('passes',       'Estilo', 'negativo — mais passes = menos lançamentos'),
    'Passes Longos':              ('passes_longos', 'Estilo', 'positivo — bola longa = mais saídas pela lateral'),
    'Passes Último Terço':        ('passes_ult3',   'Estilo', 'negativo — controlo em frente = menos lançamentos'),
    'Passes Prof. Certos':        ('passes_prof',   'Estilo', 'negativo — precisão = menos perdas'),
    'Cruzamentos':                ('cruzamentos',   'Estilo', 'positivo — jogo exterior = mais laterais'),
    'Posse de Bola %':            ('posse',         'Posse',  'negativo — mais posse = menos lançamentos'),
    'Faltas Adversário':          ('faltas_adv',    'Pressão','positivo — mais faltas adv = mais pressão = mais lançamentos'),
    'Desarmes Adversário':        ('desarmes_adv',  'Pressão','positivo — desarmes adversário aumentam lançamentos'),
    'Interceções Adversário':     ('intercecoes_adv','Pressão','positivo — interceções adv = mais perdas pela lateral'),
    'Alívios':                    ('alivios',       'Defesa', 'positivo — alívios = bola saída para lateral'),
    'Faltas Cometidas':           ('faltas',        'Defesa', 'positivo — agressividade = mais bolas no ar'),
    'Desarmes Cometidos':         ('desarmes',      'Defesa', 'positivo — desarme = ressalto lateral'),
    'Interceções':                ('intercecoes',   'Defesa', 'positivo — interceção = bola para lateral'),
    'Duelos Ganhos':              ('duelos',        'Intensidade','neutro/positivo — luta física = mais lançamentos'),
    'Foras de Jogo':              ('foras_jogo',    'Intensidade','positivo — linha alta = mais lançamentos'),
    'Livres':                     ('livres',        'Intensidade','positivo — mais faltas = mais paragens'),
    'Toques na Área Adversária':  ('toques_area',   'Ataque', 'negativo — pressão finalista = menos lançamentos'),
    'Total Remates':              ('remates',       'Ataque', 'neutro — rematar pode gerar recuperação lateral'),
}

corr_rows = []
for nome, (col, grupo, interp) in variaveis.items():
    sub = ap[['lancamentos', col]].dropna()
    if len(sub) < 3 or col not in ap.columns:
        continue
    r, p = stats.pearsonr(sub[col], sub['lancamentos'])
    sinal = '+' if r > 0 else '–'
    forca = 'forte' if abs(r) > 0.4 else 'moderada' if abs(r) > 0.2 else 'fraca'
    sig   = '***' if p < 0.001 else '**' if p < 0.01 else '*' if p < 0.05 else 'n.s.'
    corr_rows.append({
        'Variável':      nome,
        'Grupo':         grupo,
        'r Pearson':     round(r, 3),
        'p-valor':       round(p, 4),
        'Sig.':          sig,
        'Força':         forca,
        'Sinal':         sinal,
        'Interpretação': interp,
    })

df_corr = pd.DataFrame(corr_rows).sort_values('r Pearson', key=abs, ascending=False)

# ════════════════════════════════════════════════════════════════════════════════
# ANÁLISE 2: Regressão Linear Múltipla — total lançamentos por jogo
# ════════════════════════════════════════════════════════════════════════════════
print("A construir regressão múltipla...")

feat_cols = [
    'total_passes', 'total_passes_longos', 'total_cruzamentos',
    'total_faltas', 'total_alivios', 'total_desarmes',
    'total_intercecoes', 'total_duelos',
]
jogo = df[feat_cols + ['total_lanc']].dropna()
X = jogo[feat_cols].values
y = jogo['total_lanc'].values

scaler = StandardScaler()
X_sc   = scaler.fit_transform(X)

reg    = LinearRegression().fit(X_sc, y)
y_pred = reg.predict(X_sc)
r2     = r2_score(y, y_pred)

# p-values via t-test manual
n, k = X_sc.shape
resid = y - y_pred
mse   = (resid**2).sum() / (n - k - 1)
var_b = mse * np.linalg.inv(X_sc.T @ X_sc).diagonal()
se    = np.sqrt(var_b)
t_val = reg.coef_ / se
p_val = 2 * (1 - stats.t.cdf(np.abs(t_val), df=n-k-1))

reg_rows = []
for name, coef, sv, tv, pv in zip(feat_cols, reg.coef_, se, t_val, p_val):
    sig = '***' if pv<0.001 else '**' if pv<0.01 else '*' if pv<0.05 else 'n.s.'
    reg_rows.append({
        'Variável (total/jogo)': name.replace('total_','').replace('_',' ').title(),
        'Coef. Padronizado':     round(coef, 3),
        'Erro Padrão':           round(sv, 3),
        't':                     round(tv, 2),
        'p-valor':               round(pv, 4),
        'Sig.':                  sig,
        'Efeito':                '+' if coef > 0 else '–',
    })

df_reg = pd.DataFrame(reg_rows).sort_values('Coef. Padronizado', key=abs, ascending=False)
print(f"  R² = {r2:.3f}  ({r2*100:.1f}% da variância explicada)")

# ════════════════════════════════════════════════════════════════════════════════
# ANÁLISE 3: Over/Under — limiares e taxas
# ════════════════════════════════════════════════════════════════════════════════
print("A calcular Over/Under...")

thresholds = [28, 30, 32, 34, 36, 38, 40]
ou_rows = []
for thresh in thresholds:
    over  = (df['total_lanc'] > thresh).mean()
    under = (df['total_lanc'] < thresh).mean()
    exact = (df['total_lanc'] == thresh).mean()
    ou_rows.append({
        'Linha':        f'Over/Under {thresh}.5',
        'Over %':       round(over * 100, 1),
        'Under %':      round(under * 100, 1),
        'Exato %':      round(exact * 100, 1),
        'Jogos Over':   int((df['total_lanc'] > thresh).sum()),
        'Jogos Under':  int((df['total_lanc'] < thresh).sum()),
        'Total Jogos':  len(df),
    })

df_ou = pd.DataFrame(ou_rows)

# ════════════════════════════════════════════════════════════════════════════════
# ANÁLISE 4: Segmentação — Over/Under por quartil de cada variável-chave
# ════════════════════════════════════════════════════════════════════════════════
print("A segmentar por variáveis-chave...")

LINHA = 35  # limiar principal (Over/Under 35.5)
df['over35'] = (df['total_lanc'] > LINHA).astype(int)

seg_vars = {
    'Posse Diferencial': 'total_posse_diff',
    'Passes Longos (total)': 'total_passes_longos',
    'Cruzamentos (total)': 'total_cruzamentos',
    'Alívios (total)': 'total_alivios',
    'Faltas (total)': 'total_faltas',
    'Passes (total)': 'total_passes',
}

seg_rows = []
for nome, col in seg_vars.items():
    sub = df[[col, 'total_lanc', 'over35']].dropna()
    q1, q2, q3 = sub[col].quantile([0.25, 0.5, 0.75])
    for label, mask in [
        (f'Baixo (≤{q2:.0f})',  sub[col] <= q2),
        (f'Alto (>{q2:.0f})',   sub[col] >  q2),
    ]:
        grp = sub[mask]
        seg_rows.append({
            'Variável':       nome,
            'Segmento':       label,
            'N jogos':        len(grp),
            'Média Lançamentos': round(grp['total_lanc'].mean(), 1),
            f'Over {LINHA}.5 %': round(grp['over35'].mean()*100, 1),
            f'Under {LINHA}.5 %': round((1-grp['over35']).mean()*100, 1),
        })

df_seg = pd.DataFrame(seg_rows)

# ════════════════════════════════════════════════════════════════════════════════
# ANÁLISE 5: Perfis táticos por equipa
# ════════════════════════════════════════════════════════════════════════════════
print("A criar perfis táticos...")

teams_list = sorted(df['home'].dropna().unique())
perfil_rows = []
for team in teams_list:
    mh = df['home'] == team
    mf = df['away'] == team

    lanc_h  = df.loc[mh, 'Lançamentos_casa'].mean()
    lanc_f  = df.loc[mf, 'Lançamentos_fora'].mean()
    lanc_g  = pd.concat([df.loc[mh,'Lançamentos_casa'], df.loc[mf,'Lançamentos_fora']]).mean()

    passe_h = df.loc[mh, 'Passes longos_casa'].mean()
    passe_f = df.loc[mf, 'Passes longos_fora'].mean()
    passe_g = pd.concat([df.loc[mh,'Passes longos_casa'], df.loc[mf,'Passes longos_fora']]).mean()

    cruz_h  = df.loc[mh, 'Cruzamentos_casa'].mean()
    cruz_f  = df.loc[mf, 'Cruzamentos_fora'].mean()
    cruz_g  = pd.concat([df.loc[mh,'Cruzamentos_casa'], df.loc[mf,'Cruzamentos_fora']]).mean()

    posse_h = df.loc[mh, 'Posse de bola_casa'].mean()
    posse_f = df.loc[mf, 'Posse de bola_fora'].mean()
    posse_g = pd.concat([df.loc[mh,'Posse de bola_casa'], df.loc[mf,'Posse de bola_fora']]).mean()

    aliv_h  = df.loc[mh, 'Alívios_casa'].mean()
    aliv_f  = df.loc[mf, 'Alívios_fora'].mean()
    aliv_g  = pd.concat([df.loc[mh,'Alívios_casa'], df.loc[mf,'Alívios_fora']]).mean()

    # Estilo tático inferido
    if posse_g > 55:
        estilo = 'Posse Dominante'
    elif passe_g > 35:
        estilo = 'Jogo Direto/Longo'
    elif cruz_g > 15:
        estilo = 'Jogo Exterior'
    else:
        estilo = 'Equilíbrio'

    perfil_rows.append({
        'Equipa':            team,
        'Estilo Tático':     estilo,
        'Lançamentos Geral': round(lanc_g, 1),
        'Lançamentos Casa':  round(lanc_h, 1),
        'Lançamentos Fora':  round(lanc_f, 1),
        'Posse Geral %':     round(posse_g, 1),
        'Passes Longos/jogo': round(passe_g, 1),
        'Cruzamentos/jogo':  round(cruz_g, 1),
        'Alívios/jogo':      round(aliv_g, 1),
    })

df_perfil = pd.DataFrame(perfil_rows).sort_values('Lançamentos Geral', ascending=False)

# ════════════════════════════════════════════════════════════════════════════════
# GERAR EXCEL
# ════════════════════════════════════════════════════════════════════════════════
print("\nA gerar Excel de análise...")

CLR_DARK   = '1F3864'
CLR_MED    = '2E75B6'
CLR_LIGHT  = 'D6E4F0'
CLR_ACCENT = 'E2EFDA'
CLR_WARN   = 'FFEB9C'

def write_df(ws, df, start_row=1, title=None, col_widths=None):
    hdr_fill = PatternFill('solid', fgColor=CLR_DARK)
    hdr_font = Font(color='FFFFFF', bold=True, size=10)
    thin     = Side(style='thin', color='CCCCCC')
    border   = Border(bottom=thin)

    if title:
        ws.cell(row=start_row, column=1, value=title).font = Font(bold=True, size=12, color=CLR_MED)
        start_row += 1

    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=start_row, column=ci, value=col)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    for ri, (_, row) in enumerate(df.iterrows()):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=start_row+ri+1, column=ci)
            c.value = val if not (isinstance(val, float) and np.isnan(val)) else None
            c.border = border
            if (ri % 2) == 1:
                c.fill = PatternFill('solid', fgColor='F2F7FB')
            c.alignment = Alignment(horizontal='center' if ci > 1 else 'left')

    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[start_row].height = 28
    ws.freeze_panes = ws.cell(row=start_row+1, column=1)


wb = openpyxl.Workbook()

# ── Sheet 1: Correlações ──────────────────────────────────────────────────────
ws1 = wb.active; ws1.title = '1_Correlações'
write_df(ws1, df_corr, title='Correlações de Pearson — Variável vs Lançamentos por Equipa/Jogo',
         col_widths=[26, 12, 12, 10, 8, 12, 8, 45])

# ── Sheet 2: Regressão ────────────────────────────────────────────────────────
ws2 = wb.create_sheet('2_Regressão')
ws2['A1'] = f'Regressão Linear Múltipla — Total Lançamentos por Jogo   |   R² = {r2:.3f}  ({r2*100:.1f}% variância explicada)  |  n = {len(jogo)} jogos'
ws2['A1'].font = Font(bold=True, size=12, color=CLR_MED)
write_df(ws2, df_reg, start_row=3,
         col_widths=[28, 18, 14, 10, 10, 8, 10])

# ── Sheet 3: Over/Under ───────────────────────────────────────────────────────
ws3 = wb.create_sheet('3_Over_Under')
avg = df['total_lanc'].mean()
med = df['total_lanc'].median()
ws3['A1'] = f'Over/Under Lançamentos por Jogo   |   Média: {avg:.1f}  |  Mediana: {med:.0f}  |  Desvio: {df["total_lanc"].std():.1f}'
ws3['A1'].font = Font(bold=True, size=12, color=CLR_MED)
write_df(ws3, df_ou, start_row=3,
         col_widths=[18, 10, 10, 10, 12, 12, 12])

# ── Sheet 4: Segmentação ──────────────────────────────────────────────────────
ws4 = wb.create_sheet('4_Segmentação')
write_df(ws4, df_seg, title=f'Impacto das Variáveis-Chave no Over/Under {LINHA}.5 Lançamentos',
         col_widths=[26, 18, 10, 22, 14, 14])

# ── Sheet 5: Perfis ───────────────────────────────────────────────────────────
ws5 = wb.create_sheet('5_Perfis_Táticos')
write_df(ws5, df_perfil, title='Perfis Táticos por Equipa — PL 2025/26',
         col_widths=[22, 20, 18, 16, 16, 14, 18, 16, 14])

wb.save(OUT_FILE)

print(f"\nCONCLUIDO: {OUT_FILE}")
print(f"\n── TOP CORRELAÇÕES (Pearson) ──")
print(df_corr[['Variável','r Pearson','Sig.','Grupo']].head(8).to_string(index=False))
print(f"\n── REGRESSÃO (R²={r2:.3f}) ──")
print(df_reg[['Variável (total/jogo)','Coef. Padronizado','Sig.']].head(6).to_string(index=False))
print(f"\n── OVER/UNDER ──")
print(df_ou[['Linha','Over %','Under %']].to_string(index=False))
