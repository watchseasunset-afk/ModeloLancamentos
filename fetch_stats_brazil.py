"""
Fetch ALL Flashscore stats para Brasil Série A Betano 2025 — 380 jogos
Grava em: data/stats_completas_BRA_2025.xlsx
Corre: python fetch_stats_brazil.py
Duração: ~10-12 min
"""
import time, re, csv, sys, subprocess
from pathlib import Path

for pkg in ['requests', 'openpyxl']:
    try:
        __import__(pkg)
    except ImportError:
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', pkg])

import requests, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── 380 Match IDs Brasil Série A 2025 (Flashscore) ───────────────────────────
MATCH_IDS = (
    "lQQTzpqQ,YyRPIpNs,ll9GRu6C,t2UXGOif,l4kgO6aK,K8HaVN6m,xUNtF2M6,rm2PPJzP,"
    "tO58T1y0,AsMlDtiJ,2ZBCBTFE,Cj0mREE8,h0MSBxAc,8zii4o8l,EDVHknge,xGr8037D,"
    "URz3J2e2,tlka2Px1,ljAK97pR,naFyAGvA,YLYhUvRH,nPbSrvtB,tMhAgefT,Yek2eHPG,"
    "8vmysIAN,tSeKp0Bb,QgUtlZHj,p02BnMtn,vNw7ydYp,fVLgaaQi,SrK1cwe4,QknxYMRh,"
    "vaDJFuCO,IPrUZrct,UmS1SIeU,n5zoW0d5,6oFBH1sC,bNjNwhbE,Qy73JNCa,IB3eLqrm,"
    "CErzIAye,4piOQp3S,v1pSJW6r,WGbyGl67,z5a7US33,KG3ePhqL,OUXfLOQk,Iox9i8Or,"
    "WWhGS6YF,MwZPmQN7,dv6oSgXr,GKpCDfIk,GGAOhZbD,Ms9D8osp,8vqvfnSj,lEpj5Czh,"
    "Gt4a1UZG,MJQQHkKc,jaIsE7JM,xdh8a84T,fXOYFTlA,8G0i3j54,rccM6PBd,YP8tk0pf,"
    "2JuWaHbC,2yZ1drhK,Ui2kmvF6,IDrp9LyQ,YwlF3aam,n9AYjMFs,250coIqJ,dKc4s9ro,"
    "fgxO1wUa,j38gV5gL,lY8SsqwR,zVtGYv0D,6kkaxN0l,v3m7z1V0,nZgPPqHr,fyFpXRO8,"
    "0CSXL1G7,t85Kq58E,hnPPNNoe,lWodnIJj,n7czHFYA,IeXE9XQp,vsfSIy4b,6HD9VEm4,"
    "8MoqFg3N,6FYkBp2k,SUMb9OW1,rF0JKcZo,Y3F1XzJi,2q7IThIG,YH8ywW2b,Y5pU5U8C,"
    "ptCtFYoQ,S2Q90flU,vTEg4xk5,ngI12GJH,IDQp6bKh,6Jrx48wP,K2KfiVgJ,AmNx8Kkt,"
    "UDmD9C9m,MVjUv14d,feZUdZPs,OfcbFst9,raKBivl2,jNSweDff,GlzXbMeS,zZG3g0Jk,"
    "OzRogiP6,lxqG24d3,ryxP0rRF,xhSj8atA,GtMnQX1e,nRmLztMU,UXBknqCM,M9FgSNKj,"
    "AVell5sA,0G2LtNZq,UBzFZ5SG,txPbwlTi,x4I7ySc4,jRONXqcT,lrxDFpyo,xIctjRCc,"
    "E3zO6GhD,rscYMSrB,MR8POlEb,tAS9UJpK,Y77lNywf,E15HQAqn,S60cLF76,Qs4tPc8s,"
    "UcetL6DN,0Kb5JgxJ,Q7mXcL9r,Uy213snL,MZttdave,GhQrGKVQ,SCL6s0P1,8SMzIt1E,"
    "SWqlfw97,8j5N8xO0,Y98h53H8,zw7FAbgl,dCmBGxiK,zqKT3a7E,ADfFhrId,ravZCC3B,"
    "0KpGdiJU,6ZlBhQX2,6cXcqMfk,2mi7f4mp,zZzREYYb,GWKPGf4n,zFtSlrXR,vmmJj42F,"
    "tQYsALBc,zDaj2ekg,4ApwevrQ,M9ZGmdz0,jkRXqEZO,4hi7bDlI,IyTPoz5C,x4hiXzjJ,"
    "lrM9kI5m,dSma0ZJ5,UFdr4HKt,rsHer1bK,txarZdL6,4WcStlNF,fLz1MJy1,EP7RxuMs,"
    "tbCXK60d,zJKaOu7l,2ov9Kc6D,844JrAh3,8O0zyJjf,hfP8ejXp,tG53HXg4,lt9BFBOG,"
    "4IIgSl8M,KrOOMSVq,8jPpUAwA,plCfJgPi,zwRxWW8c,KUrj8TGj,AecKDkgT,2c0hZm2N,"
    "x8r3lZPH,C0cwfwdt,bXmD6Gun,ALtBnDfU,ttkfjee5,YFonhHQh,zwqx1C9N,6ciL4fAb,"
    "SCuU2YvB,llsSnMlD,hMmJlrK0,E7AWcncE,Wronjzv8,WK2aavHL,jBfi10n9,0O8vdQRQ,"
    "K8kBj4kl,OIsvhd9e,O425wcf2,SjlVgIvq,WpnjY1IM,vi9HlN23,C47Pn1XF,OrRUQJAd,"
    "W6NQpu2S,ULkZypJc,QmqrzOmA,xS2mEtRj,GINMSutp,2kEduJPk,SK5TvhYo,Gxna5CrJ,"
    "lpqazXsg,tKuixgDt,2gAtIkSO,zsJPLXSa,SGFHNgbm,rqvoTFzQ,AaX7YAC5,UuCYJBcC,"
    "rRpFWlsI,dzWc7YD6,O8hHfeU0,Sta8dHal,SfvsBGEs,WADx2wxR,OMXk9fqf,UoXstdF7,"
    "MssPTRPp,KlaBk9n4,8WrZsIpe,KCsbt59j,65Nh2nW9,YiPp481c,SE2JmmHG,lUjSoRoT,"
    "EyUAiVHi,06uRqvGr,KjUwO2xP,AF9mr6Nt,EDkDx2iI,Iod4vOM5,fXdpyAXA,WUcdtphg,"
    "OOzVPO7C,O6PfOBW7,jBxNRpw0,08KFXF9L,A1UET68m,4tR6Zyu9,xzZKBYn2,422snFOE,"
    "2HKxae3d,pIVb6wBM,IP2PMbe3,n5ftJGfS,Uk7V0HYq,jw1XKxQF,zRIyxy0E,f9A0crSi,"
    "GQ5LZvzo,Ol2Hg0RG,tnnXeVjm,pQ3FwlyJ,WE5PivdT,lCkuf9La,pvqlhmjC,dEA01iTP,"
    "2ZD8eMd4,MNjPSWM0,GpCbsWyf,KIIkqh7s,UPBpzFUQ,lbA6uA66,AqXJlEqK,UmfXQAiD,"
    "8KQkEX77,QecGUhhl,KKmIHoc2,xdI2PcG8,CfpAJ7Sk,bmlZHfio,A9GANyoL,n9UTwxE4,"
    "n5xsGgwe,h6XyxGrH,jsyLubqh,vuZYHF8r,WIuDCaF5,2wldGsGt,ptnT8HEU,Ohx5ELpg,"
    "v372NGTN,CE8ndvVO,O6HwTKan,CSmLAwpI,WWKoRbUb,hn9fPxaB,hChob00C,6eU3oNG6,"
    "ChXBq1oJ,hlAb74wQ,rFKE7DX8,OW0U1r1m,Kbfw0MVa,xGxmk5Hs,j7ITqAu2,ATzemqnf,"
    "OjP6vF3M,Gl4oXWed,GQLbtyZ9,2sptgEIF,4UIBKgCj,dO5crc4c,UiUKoWAk,b9wkihmS,"
    "8O0wZhRq,UToYfzl3,G4pOTck4,CIbbDx5N,ru2NBPTI,xKYWRyKG,8lpU7GSp,Gj5kFbzB,"
    "ppxuQFlT,OdSGVJKi,8KIYItyn,UZ6sHK5b,6g1sFwJD,YBe820Z7,hGSYILJl,K0HcHTas,"
    "dps3OsdR,lIQRMkxK,bTcZGal1,nD8EDoa6,0ImfQ3RE,EaA6F7Ug,bgsQkNk2,6a4i6rzq,"
    "WINrwp6d,E1PCDQs9,S47oc5cS,nFNKB4CL,OCAU0mb3,Kv4waRSF,CxqIiqLk,Cd7a4M4e,"
    "plCSpy85,GUen1dhn,tG5yqFxI,j9Fzu6yp,OfafazNb,ET3qsg7U,SCu2UFNO,G8KBlJ8t,"
    "pAm3cEiB,xO8Kncwg,IaDCJb96,25fxzKgm,W2uPS2nK,lKLeNtAs,4dmw5afD,QPGKHxvJ,"
    "bXmoYcOa,6Lg7utWP,WpngWygC,MXJ3LKvf"
).split(',')

NINJA_URL = 'https://global.flashscore.ninja/20/x/feed/df_st_1_{mid}'
NINJA_HDR = {
    'X-Fsign': 'SW9D1eZo',
    'Referer': 'https://www.flashscore.com/',
    'Accept': '*/*',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
}
DELAY    = 1.5
DATA_DIR = Path(__file__).parent / 'data'
RAW_CSV  = DATA_DIR / 'stats_raw_brazil.csv'
OUT_XLSX = DATA_DIR / 'stats_completas_BRA_2025.xlsx'
DATA_DIR.mkdir(exist_ok=True)

STAT_RE = re.compile(r'SG÷([^¬]+)¬SH÷([^¬]*)¬SI÷([^¬~]*)')

session = requests.Session()
session.headers.update(NINJA_HDR)

# ── Carregar progresso ─────────────────────────────────────────────────────────
done = set()
all_rows = []
if RAW_CSV.exists():
    with open(RAW_CSV, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            done.add(row['mid'])
            all_rows.append(dict(row))
    print(f'[resume] {len(done)} já recolhidos')

all_cols_seen = set()
for r in all_rows:
    all_cols_seen.update(r.keys())

# ── Scraping ───────────────────────────────────────────────────────────────────
todo = [m for m in MATCH_IDS if m not in done]
print(f'{len(todo)} jogos por recolher...\n')

with open(RAW_CSV, 'a', newline='', encoding='utf-8-sig') as f:
    writer = None

    for i, mid in enumerate(todo, 1):
        print(f'[{i}/{len(todo)}] {mid}', end=' ')
        try:
            r = session.get(NINJA_URL.format(mid=mid), timeout=15)
            r.raise_for_status()
            text = r.content.decode('utf-8', errors='replace')

            row = {'mid': mid}
            seen_stats = set()
            for m in STAT_RE.finditer(text):
                name = m.group(1).strip()
                if name not in seen_stats:
                    seen_stats.add(name)
                    row[name + '_casa'] = m.group(2).strip()
                    row[name + '_fora'] = m.group(3).strip()
                    all_cols_seen.add(name + '_casa')
                    all_cols_seen.add(name + '_fora')

            all_rows.append(row)

            if writer is None:
                cols = ['mid'] + sorted(c for c in all_cols_seen if c != 'mid')
                writer = csv.DictWriter(f, fieldnames=cols, extrasaction='ignore')
                if not done:
                    writer.writeheader()

            writer.writerow(row)
            f.flush()
            print(f"OK ({len(seen_stats)} stats)")
        except Exception as e:
            print(f'ERRO: {e}')
            all_rows.append({'mid': mid})

        time.sleep(DELAY)

# ── Excel ──────────────────────────────────────────────────────────────────────
print('\nA gerar Excel...')
all_cols = ['mid'] + sorted(c for c in all_cols_seen if c != 'mid')

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Stats BRA 2025'

hdr_fill = PatternFill('solid', fgColor='009C3B')   # verde Brasil
hdr_font = Font(color='FFDF00', bold=True)           # amarelo Brasil
for ci, col in enumerate(all_cols, 1):
    cell = ws.cell(row=1, column=ci, value=col)
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal='center')

for ri, row in enumerate(all_rows, 2):
    vals = []
    for col in all_cols:
        val = row.get(col, '')
        try:
            val = float(val.replace('%','').strip()) if val else None
        except Exception:
            pass
        vals.append(val)
    ws.append(vals)

for ci in range(1, len(all_cols)+1):
    ws.column_dimensions[get_column_letter(ci)].width = 18
ws.column_dimensions['A'].width = 12
ws.freeze_panes = 'B2'

wb.save(OUT_XLSX)
print(f'\n✓ CONCLUÍDO: {OUT_XLSX}')
print(f'  {len(all_rows)} jogos | {len(all_cols)} colunas')
