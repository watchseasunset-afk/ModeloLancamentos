"""
Calcula médias por equipa (casa, fora, geral) para todas as estatísticas PL 2025/26.
Cria novo Excel com sheets separadas.

Corre: python calc_medias_equipas.py
"""
import sys, subprocess
for pkg in ['pandas', 'openpyxl']:
    try:
        __import__(pkg)
    except ImportError:
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', pkg,
                               '--break-system-packages'])

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

DATA_DIR = Path(__file__).parent / 'data'
STATS_FILE = DATA_DIR / 'stats_completas_PL_2025_26.xlsx'
TEAMS_MAP  = Path(__file__).parent / 'teams_map.csv'
OUT_FILE   = DATA_DIR / 'medias_equipas_PL_2025_26.xlsx'

# ── Carregar dados ─────────────────────────────────────────────────────────────
print("A carregar stats...")
stats = pd.read_excel(STATS_FILE)
teams = pd.read_csv(TEAMS_MAP)

print(f"  Stats: {len(stats)} jogos, {len(stats.columns)} colunas")
print(f"  Teams: {len(teams)} entradas")

# Merge
df = stats.merge(teams, on='mid', how='left')
print(f"  Merged: {len(df)} jogos")

# Identificar colunas de stats (excluindo mid, home, away)
stat_cols = [c for c in stats.columns if c != 'mid']

# Colunas _casa e _fora
casa_cols = [c for c in stat_cols if c.endswith('_casa')]
fora_cols = [c for c in stat_cols if c.endswith('_fora')]

# Converter para numérico
for c in stat_cols:
    df[c] = pd.to_numeric(df[c], errors='coerce')

print(f"\n  Colunas _casa: {len(casa_cols)}")
print(f"  Colunas _fora: {len(fora_cols)}")

# ── Nomes base das estatísticas ────────────────────────────────────────────────
stat_names = sorted(set(c.replace('_casa','').replace('_fora','') for c in stat_cols))
print(f"  Stats únicas: {len(stat_names)}")

# ── Calcular médias ────────────────────────────────────────────────────────────
teams_list = sorted(df['home'].dropna().unique())
print(f"\n  Equipas encontradas: {len(teams_list)}")

records_home    = []
records_away    = []
records_overall = []

for team in teams_list:
    mask_home = df['home'] == team
    mask_away = df['away'] == team

    row_home = {'Equipa': team, 'Jogos_Casa': mask_home.sum()}
    row_away = {'Equipa': team, 'Jogos_Fora': mask_away.sum()}
    row_all  = {'Equipa': team, 'Jogos_Total': (mask_home | mask_away).sum()}

    for stat in stat_names:
        col_c = stat + '_casa'
        col_f = stat + '_fora'

        # Casa: quando jogou em casa, os seus valores estão em _casa
        if col_c in df.columns:
            row_home[f'{stat}_media'] = df.loc[mask_home, col_c].mean()

        # Fora: quando jogou fora, os seus valores estão em _fora
        if col_f in df.columns:
            row_away[f'{stat}_media'] = df.loc[mask_away, col_f].mean()

        # Geral: combinar jogos casa e fora
        vals_home = df.loc[mask_home, col_c] if col_c in df.columns else pd.Series(dtype=float)
        vals_away = df.loc[mask_away, col_f] if col_f in df.columns else pd.Series(dtype=float)
        combined  = pd.concat([vals_home, vals_away])
        row_all[f'{stat}_media'] = combined.mean()

    records_home.append(row_home)
    records_away.append(row_away)
    records_overall.append(row_all)

df_home    = pd.DataFrame(records_home)
df_away    = pd.DataFrame(records_away)
df_overall = pd.DataFrame(records_overall)

# ── Posse de bola separada (já está incluída mas destacar) ────────────────────
POSSE_COL = 'Posse de bola'
posse_records = []
for team in teams_list:
    mask_home = df['home'] == team
    mask_away = df['away'] == team
    col_c = POSSE_COL + '_casa'
    col_f = POSSE_COL + '_fora'
    posse_h = df.loc[mask_home, col_c].mean() if col_c in df.columns else None
    posse_f = df.loc[mask_away, col_f].mean() if col_f in df.columns else None
    vals    = []
    if col_c in df.columns: vals.append(df.loc[mask_home, col_c])
    if col_f in df.columns: vals.append(df.loc[mask_away, col_f])
    posse_g = pd.concat(vals).mean() if vals else None
    posse_records.append({
        'Equipa': team,
        'Posse_Casa_%': round(posse_h, 1) if posse_h is not None else None,
        'Posse_Fora_%': round(posse_f, 1) if posse_f is not None else None,
        'Posse_Geral_%': round(posse_g, 1) if posse_g is not None else None,
    })

df_posse = pd.DataFrame(posse_records).sort_values('Posse_Geral_%', ascending=False)

# ── Lançamentos separada (destaque) ───────────────────────────────────────────
LANC_COL = 'Lançamentos'
lanc_records = []
for team in teams_list:
    mask_home = df['home'] == team
    mask_away = df['away'] == team
    col_c = LANC_COL + '_casa'
    col_f = LANC_COL + '_fora'
    lanc_h = df.loc[mask_home, col_c].mean() if col_c in df.columns else None
    lanc_f = df.loc[mask_away, col_f].mean() if col_f in df.columns else None
    vals   = []
    if col_c in df.columns: vals.append(df.loc[mask_home, col_c])
    if col_f in df.columns: vals.append(df.loc[mask_away, col_f])
    lanc_g = pd.concat(vals).mean() if vals else None
    lanc_records.append({
        'Equipa': team,
        'Lançamentos_Casa': round(lanc_h, 1) if lanc_h is not None else None,
        'Lançamentos_Fora': round(lanc_f, 1) if lanc_f is not None else None,
        'Lançamentos_Geral': round(lanc_g, 1) if lanc_g is not None else None,
    })

df_lanc = pd.DataFrame(lanc_records).sort_values('Lançamentos_Geral', ascending=False)

# ── Escrever Excel ─────────────────────────────────────────────────────────────
print("\nA gerar Excel...")

def style_sheet(ws, df, title):
    """Aplica estilo ao worksheet."""
    hdr_fill = PatternFill('solid', fgColor='1F3864')
    hdr_font = Font(color='FFFFFF', bold=True, size=10)

    # Header
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Dados
    for ri, row in df.iterrows():
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri+2, column=ci)
            if isinstance(val, float):
                cell.value = round(val, 2)
            else:
                cell.value = val
            if (ri+2) % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='EEF2F7')

    # Largura colunas
    ws.column_dimensions['A'].width = 20
    for ci in range(2, len(df.columns)+1):
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.freeze_panes = 'B2'
    ws.row_dimensions[1].height = 30


wb = openpyxl.Workbook()

# Sheet 1: Posse de bola
ws1 = wb.active
ws1.title = 'Posse de Bola'
style_sheet(ws1, df_posse, 'Posse de Bola')

# Sheet 2: Lançamentos
ws2 = wb.create_sheet('Lançamentos')
style_sheet(ws2, df_lanc, 'Lançamentos Laterais')

# Sheet 3: Médias Gerais (todas as stats)
df_overall_r = df_overall.copy()
for c in df_overall_r.columns:
    if c not in ('Equipa', 'Jogos_Total'):
        df_overall_r[c] = df_overall_r[c].round(2)
df_overall_r = df_overall_r.sort_values('Equipa')
ws3 = wb.create_sheet('Médias Gerais')
style_sheet(ws3, df_overall_r, 'Médias Gerais')

# Sheet 4: Médias Casa
df_home_r = df_home.copy()
for c in df_home_r.columns:
    if c not in ('Equipa', 'Jogos_Casa'):
        df_home_r[c] = df_home_r[c].round(2)
df_home_r = df_home_r.sort_values('Equipa')
ws4 = wb.create_sheet('Médias Casa')
style_sheet(ws4, df_home_r, 'Médias Casa')

# Sheet 5: Médias Fora
df_away_r = df_away.copy()
for c in df_away_r.columns:
    if c not in ('Equipa', 'Jogos_Fora'):
        df_away_r[c] = df_away_r[c].round(2)
df_away_r = df_away_r.sort_values('Equipa')
ws5 = wb.create_sheet('Médias Fora')
style_sheet(ws5, df_away_r, 'Médias Fora')

wb.save(OUT_FILE)
print(f"\nCONCLUIDO: {OUT_FILE}")
print(f"  Sheets: Posse de Bola | Lançamentos | Médias Gerais | Médias Casa | Médias Fora")
print(f"  Equipas: {len(teams_list)}")
print(f"\nTop 5 Lançamentos (geral):")
print(df_lanc[['Equipa','Lançamentos_Geral']].head(5).to_string(index=False))
print(f"\nTop 5 Posse (geral):")
print(df_posse[['Equipa','Posse_Geral_%']].head(5).to_string(index=False))
