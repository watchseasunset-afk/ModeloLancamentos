"""
Fetch lançamentos do Marítimo + Académico Viseu (Liga Portugal 2, época 2024/25)
e Vitória Guimarães jogos em CASA (PPL 2025/26, IDs em falta no teams_map).

Corre localmente:  python fetch_maritimo_liga2.py
Resultado:  imprime as médias e atualiza medias_equipas_MULTI_LIGA.xlsx

Porquê este script?
- Marítimo e AVS vieram da Liga 2 → sem dados no stats_completas_PPL_2025_26.xlsx
- Vitória Guimarães → bug no teams_map_ppl.csv (só aparece como 'away')
"""
import re, time, json, sys, urllib.request, urllib.error
from pathlib import Path

# ─── Config ────────────────────────────────────────────────────────────────────
DATA_DIR = Path(__file__).parent / 'data'
XLSX_OUT = DATA_DIR / 'medias_equipas_MULTI_LIGA.xlsx'
DELAY    = 1.5

NINJA_URL  = 'https://global.flashscore.ninja/20/x/feed/df_st_1_{mid}'
FS_HDR = {
    'X-Fsign':    'SW9D1eZo',
    'Referer':    'https://www.flashscore.com/',
    'Accept':     '*/*',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
}

# Liga Portugal 2, época 2024/25
# URL para scraping da página de resultados (Googlebot UA)
LP2_URL  = 'https://www.flashscore.com/football/portugal/liga-portugal-2-2024-2025/results/'
LP2B_URL = 'https://www.flashscore.pt/futebol/portugal/liga-portugal-2-2024-2025/resultados/'

# PPL 2025/26 — para jogos em CASA do Vitória Guimarães
PPL_URL  = 'https://www.flashscore.com/football/portugal/liga-portugal-betclic/results/'

GBOT_UA  = 'Mozilla/5.0 (compatible; Googlebot/2.1; +http://www.google.com/bot.html)'

# Regex para extrair match IDs da página Flashscore
MID_RE    = re.compile(r'~([A-Za-z0-9]{8})~')
STAT_RE   = re.compile(r'SG÷([^¬]+)¬SH÷([^¬]*)¬SI÷([^¬~]*)')
HOME_RE   = re.compile(r'CX÷([^¬~]+)')
AWAY_RE   = re.compile(r'AF÷([^¬~]+)')

# ─── Targets ───────────────────────────────────────────────────────────────────
# Equipas que queremos (normalizado para substring match)
TARGETS = {
    'maritimo':  'Maritimo',
    'academico': 'Academico Viseu',
    'guimaraes': 'Vitoria Guimaraes',
}

# ─── Funções ───────────────────────────────────────────────────────────────────
def _req(url, ua=None):
    hdr = dict(FS_HDR)
    if ua: hdr['User-Agent'] = ua
    req = urllib.request.Request(url, headers=hdr)
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            return r.read().decode('utf-8', errors='replace')
    except Exception as e:
        print(f'  ERRO fetch {url}: {e}')
        return ''

def get_match_ids_with_teams(league_url):
    """Scrape página Flashscore e devolve lista de (mid, home, away)."""
    print(f'  A carregar {league_url}...')
    html = _req(league_url, ua=GBOT_UA)
    if not html or len(html) < 3000:
        print(f'  Resposta insuficiente ({len(html)} bytes). Tentar URL alternativa.')
        return []

    # Flashscore inclui dados de jogos no HTML
    # Padrão: ~MIDXXXXX~...CX÷HomeTeam¬...AF÷AwayTeam¬
    match_blocks = re.findall(r'~([A-Za-z0-9]{8})~[^~]{0,500}?CX÷([^¬]+)¬[^~]*?AF÷([^¬]+)¬', html)

    results = []
    for mid, home, away in match_blocks:
        results.append((mid.strip(), home.strip(), away.strip()))

    if not results:
        # Fallback: só IDs (sem teams info no HTML)
        mids = list(dict.fromkeys(MID_RE.findall(html)))  # unique, ordered
        print(f'  {len(mids)} match IDs encontrados (sem teams info)')
        return [(m, '', '') for m in mids]

    print(f'  {len(results)} jogos encontrados')
    return results

def fetch_stats(mid):
    """Devolve dict {stat_name: (casa_val, fora_val)} para um match."""
    text = _req(NINJA_URL.format(mid=mid))
    if not text: return {}
    stats = {}
    for m in STAT_RE.finditer(text):
        name = m.group(1).strip()
        if name not in stats:  # primeira ocorrência = total jogo
            try: casa = float(m.group(2).strip())
            except: casa = None
            try: fora = float(m.group(3).strip())
            except: fora = None
            stats[name] = (casa, fora)
    return stats

def contains_team(name, team_key):
    return team_key.lower() in name.lower()

# ─── Main ──────────────────────────────────────────────────────────────────────
print('='*60)
print('Fetch lançamentos: Marítimo, AVS, Vitória Guimarães')
print('='*60)

# ── 1. Liga Portugal 2, 2024/25 ─────────────────────────────────────────────
print('\n[1] Liga Portugal 2, 2024/25...')
lp2_games = get_match_ids_with_teams(LP2_URL)
if not lp2_games:
    lp2_games = get_match_ids_with_teams(LP2B_URL)

# ── 2. PPL 2025/26 (para Guimarães em casa) ──────────────────────────────────
print('\n[2] PPL 2025/26...')
ppl_games = get_match_ids_with_teams(PPL_URL)

# ── 3. Separar jogos por equipa ───────────────────────────────────────────────
team_games = {
    'maritimo':  {'home': [], 'away': []},
    'academico': {'home': [], 'away': []},
    'guimaraes': {'home': [], 'away': []},
}

def classify_games(games_list, target_key):
    """Filtra e classifica jogos home/away para uma equipa."""
    for mid, home, away in games_list:
        if contains_team(home, target_key):
            team_games[target_key]['home'].append(mid)
        elif contains_team(away, target_key):
            team_games[target_key]['away'].append(mid)

for key in ['maritimo', 'academico']:
    classify_games(lp2_games, key)

classify_games(ppl_games, 'guimaraes')

for key, info in team_games.items():
    print(f'  {key}: {len(info["home"])} home, {len(info["away"])} away games found')

# ── 4. Fetch stats e calcular médias ─────────────────────────────────────────
print('\n[3] A recolher stats (ninja API)...')

results = {}

for team_key, sides in team_games.items():
    home_lanc = []
    away_lanc = []

    # Home games: usar Lançamentos_casa
    for mid in sides['home']:
        stats = fetch_stats(mid)
        lanc = stats.get('Lançamentos', (None, None))[0]
        if lanc is not None:
            home_lanc.append(lanc)
            print(f'  {team_key} HOME {mid}: lanc_casa={lanc}')
        time.sleep(DELAY)

    # Away games: usar Lançamentos_fora
    for mid in sides['away']:
        stats = fetch_stats(mid)
        lanc = stats.get('Lançamentos', (None, None))[1]
        if lanc is not None:
            away_lanc.append(lanc)
            print(f'  {team_key} AWAY {mid}: lanc_fora={lanc}')
        time.sleep(DELAY)

    all_lanc = home_lanc + away_lanc
    avg = round(sum(all_lanc) / len(all_lanc), 2) if all_lanc else None
    results[team_key] = {
        'home_lanc': home_lanc,
        'away_lanc': away_lanc,
        'avg': avg,
        'n': len(all_lanc),
    }
    print(f'\n  {TARGETS[team_key]}: avg={avg} ({len(all_lanc)} jogos)')

print('\n[4] Resultados finais:')
print('-'*40)
for key, res in results.items():
    print(f'{TARGETS[key]:25s}: lanc_media={res["avg"]}  ({res["n"]} jogos)')

# ── 5. Atualizar Excel ────────────────────────────────────────────────────────
print('\n[5] A atualizar medias_equipas_MULTI_LIGA.xlsx...')

try:
    import openpyxl
    from openpyxl.styles import PatternFill

    wb = openpyxl.load_workbook(XLSX_OUT)
    ws = wb['PPL - Geral']

    headers = {cell.value: cell.column for cell in ws[1]}
    equipa_col = headers.get('Equipa')
    liga_col   = headers.get('Liga')
    jogos_col  = headers.get('Jogos_Total')
    lanc_col   = headers.get('Lançamentos_media')

    if not all([equipa_col, lanc_col]):
        print('  ERRO: colunas não encontradas no Excel')
    else:
        # Verificar equipas existentes
        existing = {}
        for row in ws.iter_rows(min_row=2, values_only=False):
            eq = row[equipa_col-1].value
            if eq:
                existing[eq] = row[equipa_col-1].row

        update_map = {
            'Maritimo':         results.get('maritimo', {}).get('avg'),
            'Academico Viseu':  results.get('academico', {}).get('avg'),
            'Vitoria Guimaraes': results.get('guimaraes', {}).get('avg'),
        }

        for eq_name, avg in update_map.items():
            if avg is None:
                print(f'  {eq_name}: sem dados — a saltar')
                continue
            if eq_name in existing:
                row_num = existing[eq_name]
                ws.cell(row=row_num, column=lanc_col, value=avg)
                print(f'  Atualizado {eq_name}: {avg}')
            else:
                new_row = ws.max_row + 1
                ws.cell(row=new_row, column=liga_col or 1, value='PPL')
                ws.cell(row=new_row, column=equipa_col, value=eq_name)
                n = results.get(list(TARGETS.keys())[list(TARGETS.values()).index(eq_name)] if eq_name in TARGETS.values() else '', {}).get('n', 0)
                ws.cell(row=new_row, column=jogos_col, value=n)
                ws.cell(row=new_row, column=lanc_col, value=avg)
                if new_row % 2 == 0:
                    fill = PatternFill('solid', fgColor='EEF2F7')
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=new_row, column=col).fill = fill
                print(f'  Adicionado {eq_name}: {avg}')

        wb.save(XLSX_OUT)
        print(f'  Gravado: {XLSX_OUT}')

except Exception as e:
    print(f'  ERRO ao atualizar Excel: {e}')
    import traceback; traceback.print_exc()

print('\nDone. Faz railway up para fazer deploy das alterações ao app.py')
