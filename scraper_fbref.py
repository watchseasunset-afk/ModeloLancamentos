"""
Scraper FBref — Lançamentos Laterais Premier League 2025/26
============================================================
Busca todos os jogos da PL 2025/26 e extrai:
  - Data, Jornada, Casa, Fora, Golos Casa, Golos Fora
  - Lançamentos Laterais Casa, Lançamentos Laterais Fora

Grava progresso em data/lancamentos_raw.csv (resume se interrompido).
No final gera data/lancamentos_PL_2025_26.xlsx

Corre: python scraper_fbref.py
Duração estimada: ~40-50 min (380 jogos × 5s delay)
"""

import re, time, csv, os, html, subprocess, sys
from pathlib import Path

# Instalar requests se necessário
try:
    import requests
except ImportError:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'requests'])
    import requests

# ── Configuração ──────────────────────────────────────────────────────────────
FIXTURES_URL = (
    'https://fbref.com/en/comps/9/2025-2026/schedule/'
    '2025-2026-Premier-League-Scores-and-Fixtures'
)
DELAY        = 6          # segundos entre pedidos (não reduzir — FBref bloqueia)
TIMEOUT      = 25
DATA_DIR     = Path(__file__).parent / 'data'
RAW_CSV      = DATA_DIR / 'lancamentos_raw.csv'
OUTPUT_XLSX  = DATA_DIR / 'lancamentos_PL_2025_26.xlsx'

DATA_DIR.mkdir(exist_ok=True)

# Sessão com headers realistas
SESSION = requests.Session()
SESSION.headers.update({
    'User-Agent': (
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/125.0.0.0 Safari/537.36'
    ),
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.9',
    'Accept-Encoding': 'gzip, deflate, br',
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1',
    'Sec-Fetch-Dest': 'document',
    'Sec-Fetch-Mode': 'navigate',
    'Sec-Fetch-Site': 'same-origin',
    'Referer': 'https://fbref.com/',
})

# ── Helpers ───────────────────────────────────────────────────────────────────
def fetch(url):
    # Primeiro pedido sem Referer para simular entrada directa
    if 'schedule' in url:
        SESSION.headers.update({'Referer': 'https://www.google.com/'})
    else:
        SESSION.headers.update({'Referer': FIXTURES_URL})
    r = SESSION.get(url, timeout=TIMEOUT)
    r.raise_for_status()
    return r.text


def parse_fixtures(html_text):
    """Extrai lista de URLs de match reports da página de fixtures."""
    pattern = re.compile(
        r'href="(/en/matches/[a-f0-9]{8}/[^"]*Premier-League[^"]*)"'
    )
    urls = []
    seen = set()
    for m in pattern.finditer(html_text):
        path = m.group(1)
        if path not in seen and 'schedule' not in path:
            seen.add(path)
            urls.append('https://fbref.com' + path)
    return urls


def parse_match(url, html_text):
    """
    Extrai dados de um jogo: equipas, golos, lançamentos laterais.
    Retorna dict ou None se o jogo ainda não foi disputado.
    """
    # Título: "Arsenal vs. Chelsea Match Report – Saturday August 17, 2025"
    title_m = re.search(r'<title>([^<]+)</title>', html_text)
    title   = html.unescape(title_m.group(1)) if title_m else ''

    # Equipas do URL: .../Arsenal-Chelsea-August-17-2025-Premier-League
    slug_m = re.search(r'/matches/[a-f0-9]{8}/([^/]+)-Premier-League', url)
    if not slug_m:
        return None
    parts = slug_m.group(1).split('-')

    # Data (formato: Month-DD-YYYY no final do slug)
    months = {'January':1,'February':2,'March':3,'April':4,'May':5,'June':6,
              'July':7,'August':8,'September':9,'October':10,'November':11,'December':12}
    date_str = None
    for i, p in enumerate(parts):
        if p in months and i+2 < len(parts):
            try:
                date_str = f"{parts[i+2]}-{months[p]:02d}-{int(parts[i+1]):02d}"
            except Exception:
                pass
            break

    # Golos — procura no score box
    score_m = re.search(
        r'class="score"[^>]*>\s*(\d+)\s*</div>.*?class="score"[^>]*>\s*(\d+)\s*</div>',
        html_text, re.DOTALL
    )
    if not score_m:
        return None   # jogo ainda não disputado

    goals_home = int(score_m.group(1))
    goals_away = int(score_m.group(2))

    # Equipas — do score box
    team_m = re.findall(r'itemprop="name"[^>]*>\s*([^<]+)\s*</span>', html_text)
    home_team = team_m[0].strip() if len(team_m) > 0 else ''
    away_team = team_m[1].strip() if len(team_m) > 1 else ''

    # Throw Ins — padrão: número \n Throw Ins \n número
    throw_m = re.search(
        r'(\d+)\s*\n\s*Throw Ins\s*\n\s*(\d+)',
        html_text
    )
    if not throw_m:
        # Tenta variante HTML
        throw_m = re.search(
            r'>(\d+)<[^>]*>\s*Throw Ins\s*<[^>]*>(\d+)<',
            html_text
        )
    throw_home = int(throw_m.group(1)) if throw_m else None
    throw_away = int(throw_m.group(2)) if throw_m else None

    # Jornada — procura "Matchweek X"
    week_m = re.search(r'Matchweek\s+(\d+)', html_text, re.IGNORECASE)
    matchweek = int(week_m.group(1)) if week_m else None

    return {
        'date':        date_str,
        'matchweek':   matchweek,
        'home':        home_team,
        'away':        away_team,
        'goals_home':  goals_home,
        'goals_away':  goals_away,
        'throw_home':  throw_home,
        'throw_away':  throw_away,
        'url':         url,
    }


# ── Carregar progresso anterior ───────────────────────────────────────────────
done_urls = set()
rows = []
if RAW_CSV.exists():
    with open(RAW_CSV, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
            done_urls.add(row['url'])
    print(f"[resume] {len(done_urls)} jogos já recolhidos.")


# ── Passo 1: Buscar lista de fixtures ─────────────────────────────────────────
print("A buscar fixtures PL 2025/26...")
fixtures_html = fetch(FIXTURES_URL)
all_urls = parse_fixtures(fixtures_html)
print(f"  → {len(all_urls)} match reports encontrados.")

todo = [u for u in all_urls if u not in done_urls]
print(f"  → {len(todo)} por recolher.\n")

# ── Passo 2: Scraping jogo a jogo ─────────────────────────────────────────────
csv_fields = ['date','matchweek','home','away','goals_home','goals_away',
              'throw_home','throw_away','url']

with open(RAW_CSV, 'a', newline='', encoding='utf-8') as f:
    writer = csv.DictWriter(f, fieldnames=csv_fields)
    if not done_urls:
        writer.writeheader()

    for i, url in enumerate(todo, 1):
        print(f"[{i}/{len(todo)}] {url.split('/')[-1][:60]}")
        try:
            match_html = fetch(url)
            data = parse_match(url, match_html)
            if data:
                writer.writerow(data)
                f.flush()
                rows.append(data)
                ti = f"throw {data['throw_home']}/{data['throw_away']}" \
                     if data['throw_home'] is not None else 'throw N/A'
                print(f"       {data['home']} {data['goals_home']}-"
                      f"{data['goals_away']} {data['away']}  |  {ti}")
            else:
                print(f"       [sem resultado — jogo futuro ou parse falhou]")
        except Exception as e:
            print(f"       ERRO: {e}")

        time.sleep(DELAY)


# ── Passo 3: Exportar para Excel ──────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Lançamentos PL 2025-26'

    col_headers = [
        'Data', 'Jornada', 'Casa', 'Fora',
        'Golos Casa', 'Golos Fora',
        'Lanç. Casa', 'Lanç. Fora', 'Total Lanç.', 'URL'
    ]

    # Header style
    header_fill = PatternFill('solid', fgColor='1F3864')
    header_font = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(col_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Dados
    sorted_rows = sorted(rows, key=lambda r: (r.get('date') or '', r.get('matchweek') or 0))
    for row_i, r in enumerate(sorted_rows, 2):
        th = int(r['throw_home']) if r.get('throw_home') not in (None, '', 'None') else None
        ta = int(r['throw_away']) if r.get('throw_away') not in (None, '', 'None') else None
        total = (th + ta) if (th is not None and ta is not None) else None
        ws.append([
            r.get('date'), r.get('matchweek'), r.get('home'), r.get('away'),
            r.get('goals_home'), r.get('goals_away'),
            th, ta, total, r.get('url')
        ])
        # Zebra
        if row_i % 2 == 0:
            fill = PatternFill('solid', fgColor='EEF2F7')
            for col in range(1, len(col_headers)+1):
                ws.cell(row=row_i, column=col).fill = fill

    # Largura colunas
    widths = [12, 9, 22, 22, 11, 11, 11, 11, 12, 60]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = 'A2'
    wb.save(OUTPUT_XLSX)
    print(f"\n✓ Excel guardado: {OUTPUT_XLSX}")
    print(f"  Total jogos: {len(sorted_rows)}")
    with_throw = sum(1 for r in sorted_rows if r.get('throw_home') not in (None, '', 'None'))
    print(f"  Com throw-ins: {with_throw}")

except ImportError:
    print("\nopenpyxl não instalado. A instalar...")
    import subprocess
    subprocess.run(['pip', 'install', 'openpyxl'], check=True)
    print("Corre o script novamente para gerar o Excel.")
