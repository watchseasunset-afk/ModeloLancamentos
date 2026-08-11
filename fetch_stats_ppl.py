"""
Fetch ALL Flashscore stats para PPL (Liga Portugal) 2025/26 — 308 jogos
Grava em: data/stats_completas_PPL_2025_26.xlsx
Corre: python fetch_stats_ppl.py
Duração: ~8-10 min
"""
import time, re, csv, sys, subprocess
from pathlib import Path

# Instalar dependências
for pkg in ['requests', 'openpyxl']:
    try:
        __import__(pkg)
    except ImportError:
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', pkg])

import requests, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── 308 Match IDs PPL 2025/26 (Flashscore) ────────────────────────────────────
MATCH_IDS = (
    "IcJBYh6e,ppL3zFyq,f5SwUIIi,lS5CDcYA,rRbLBy3N,A1I2OEnT,UoaZIhAj,OpKfQzIG,"
    "2VVoSdm4,vcpi6FQp,t224FJ3b,KnEeHuZo,fVBU2tKh,676hNSk1,OA3oabJH,UNYocxlU,"
    "C490L6KD,AuDqPlLl,jc5w1Kl5,n58M42kt,jVdAAk6e,OKCVHXDL,E9RDUAcR,CpoJ8Tz8,"
    "j7ENJgq9,r5g2CByq,IsY4WWSE,jHHydji2,riETcCMk,Ya1ycHJ5,URCJWZNj,CnnNjeFM,"
    "vcW1qD7d,jXcEhHpA,rVxgoZxp,bJxc3YbS,jgZl5fUF,pze6fwGc,WjJm7Ga3,n3qAtdZO,"
    "fTUmtmem,bujnkrvI,2Ddwi495,I7FDzqPO,UoqWhQvg,8WfNfnAt,MuM4x5fC,lSNdvRPa,"
    "OpvIo9XP,re4uFAdl,dIAeBSeD,rk1VN5uJ,OAN8KWYQ,r9qET9Bs,OO2sZVmK,zPcNPRA6,"
    "z3MnDlR0,GMsMRmtf,lYr60Bse,zksEbkC7,QiZGM1bc,SIfc2XCr,YcnVbb6j,Gt6b4KrT,"
    "MJSPKuT9,jcP0Mh4E,8G2j6tEG,vyBZ9OEi,KGPI9f3B,CjMABGYb,6DaqeelI,n1Es82q4,"
    "CE4c2Ceo,GrJR7YXN,jyY1Dx4n,2ughgZIU,vwt2rI4C,WEHD48LE,jH9K0aKt,48PRt5T7,"
    "xY7Sbwkg,Aclon05m,rBngpvz0,dMpqWmFk,h8Uoj96L,rgyJrRbe,4nliURq2,6i64wWMF,"
    "zTXApmUr,APCeuhh3,KfWwhVy9,Cl9CyAiS,QJVVOW7M,tWeS4BGj,IVP9Hl0d,dvvFSE8c,"
    "j1HmCgOG,44sNQhwA,M3W0JAVq,ryEuEFg4,p0AdAXhT,4CFwb336,xvgKAtmC,dKXpdsYI,"
    "MLPWFyPi,8pMUaPYg,fFXphYWp,8b13EOmm,4QfCC2Ia,xGAXJpQP,tA2mmnkl,MPLGh8RQ,"
    "Gnsj77tK,zXNM1o4s,lz6eoQJ0,QwjGlv9e,lCN8fUdE,Y3gOnIw8,Akf4q4lD,YuBydMGL,"
    "jmb5zllo,Chq0dVA4,4O17j0vq,6ethbith,SSuybrn9,zZhkzuf2,v7atx1Pk,Y7X8f9uH,"
    "jJW9oEdn,0fi0RABU,4SciTWtI,rulTYzDt,WEaqVhC5,CYKIqhRb,j5iyXEsg,bkJQsWeB,"
    "fHCZuAQN,Wvqz4GD6,QstS5xqf,vRNIEXRO,W0KAGgcC,K0PUPczQ,vo61Kybm,Cb49IFSa,"
    "hfXr2frJ,UFnJ7bEs,IekDB7FF,bTPoQ6U8,zyG9inOq,rLsa23x2,UZ8IkQhd,MTe4DTp3,"
    "QLhL9opS,AgUi4P7k,4hPwc80M,WIZWkEVs,zwHDZD2b,O2k1Q9Pp,d6penin4,dUSswm9j,"
    "4I8UVV1N,4Wm3pVGG,rwyVbUV9,zDXM0j1c,0bsCr9oT,zidnlDHi,4WXvlh0f,lMI5yzYo,"
    "W4AMXiXA,E1g0ez8K,ne1EXWFD,8OUepAaJ,0bRmnWU6,GWhRrcOE,0OAvLgW7,WrJNOyXr,"
    "IF9cyFGl,lpiZtygR,61ZrWJv9,2TSiUc9L,hb7WMF1e,QuD5Zho1,niJ9HKPF,j5Z6caYq,"
    "CM2ZF7lB,867IFbfS,IuglAxIk,vRF1Jte3,OIXEew3d,E7nc8Gn2,4daRHTJb,02Ha3U45,"
    "hIV8an4U,hdfYn9KO,YFF718ZH,naiQlVkC,rsBIJkkn,jXcHjiLa,dAKr7C5t,4zOj5jzg,"
    "Ora1PEM0,tKVRuz7s,xnPzvEyf,bZIxbZSP,W829NhiD,IJa9hDjm,pCSrxh66,6RGizWzJ,"
    "plpdodqK,xSSJo5w3,pYuYN1od,INegVqgM,nPKhRFzi,40kwZRgc,l6nPPNGq,2TAhRzhl,"
    "d87pX5O9,dUrLqq8F,8As6zKD4,fRuEYcrH,SjMWLKbB,QRPFP2an,jgZMWyDT,YiTNNtUb,"
    "04KvKbTN,negcxtqh,Kl1BExio,veZuDPFt,zNSlB3pg,MDa55aEU,SIpsFCoR,EFIU588D,"
    "Emdc7LqI,n3RD9j9l,8zRd9sF5,h0UL7Uw1,f5nZGYGE,vc4ngVO7,4GXj4eX8,CMp5thu2,"
    "ht6veife,0UoRmzmS,0rRa2Z1L,WQHVdDPr,WCxcrEAk,WSIZvu4c,hr7rwJY9,KhiNVzRq,"
    "6kRmHyCj,Cl9FZqSH,WlfAiIm3,08Ajyc3M,vFcIkdIF,h6kVTEed,phKn4iwn,r9HbwSSh,"
    "C62oWALP,UNrIQlqm,2ysQOSDa,vyB7y6c5,p0Aa87TI,6TwsLpDO,b3zZM6rC,As4iulbt,"
    "0nHBJXyK,EVfAF3Jl,2wdWMoRE,zoqJDsl1,OpMjATb6,GKIrCkUg,rBnRBLID,b5kuLPeR,"
    "dI2sEBas,bNfJnEqo,Aq9parZ7,0n1mNfLc,l8XU4D5T,8C4eLYkA,EHFy144e,nqyE8e64,"
    "S0HT2QZr,f9vM6ZzH,pvJ0aiTj,UixehzMN,0CpvdIMb,dvvmfdiB,lYGLwy7U,ryEDucxI,"
    "hnsWcvhn,n9K4sJ75,QTn5AHyh,hKpskA9E,lEhvRcNO,hUmeR6H7,O67EW1gm,KY0NUuOa,"
    "EeMdquwg,Sb4VSJhC,ITSmo18t,p4pmTSne,l2Sw2MDF,YRV209xB,tvfxplhU,4Ehzr1Mq,"
    "SnvV3rq3,hINf2V7b,0tnqtuid,YDjUoANH"
).split(',')

NINJA_URL = 'https://global.flashscore.ninja/20/x/feed/df_st_1_{mid}'
NINJA_HDR = {
    'X-Fsign': 'SW9D1eZo',
    'Referer': 'https://www.flashscore.com/',
    'Accept': '*/*',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
}
DELAY    = 1.5
DATA_DIR = Path(__file__).parent / 'data'
RAW_CSV  = DATA_DIR / 'stats_raw_ppl.csv'
OUT_XLSX = DATA_DIR / 'stats_completas_PPL_2025_26.xlsx'
DATA_DIR.mkdir(exist_ok=True)

STAT_RE = re.compile(r'SG÷([^¬]+)¬SH÷([^¬]*)¬SI÷([^¬~]*)')

session = requests.Session()
session.headers.update(NINJA_HDR)

# ── Carregar progresso ─────────────────────────────────────────────────────────
done = set()
all_rows = []
if RAW_CSV.exists():
    with open(RAW_CSV, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            done.add(row['mid'])
            all_rows.append(dict(row))
    print(f'[resume] {len(done)} já recolhidos')

all_cols_seen = set()
for r in all_rows:
    all_cols_seen.update(r.keys())

# ── Scraping ───────────────────────────────────────────────────────────────────
todo = [m for m in MATCH_IDS if m not in done]
print(f'{len(todo)} jogos por recolher...\n')

with open(RAW_CSV, 'a', newline='', encoding='utf-8-sig') as f:
    writer = None

    for i, mid in enumerate(todo, 1):
        print(f'[{i}/{len(todo)}] {mid}', end=' ')
        try:
            r = session.get(NINJA_URL.format(mid=mid), timeout=15)
            r.raise_for_status()
            text = r.content.decode('utf-8', errors='replace')

            row = {'mid': mid}
            seen_stats = set()
            for m in STAT_RE.finditer(text):
                name = m.group(1).strip()
                if name not in seen_stats:  # primeira ocorrência = total jogo
                    seen_stats.add(name)
                    row[name + '_casa'] = m.group(2).strip()
                    row[name + '_fora'] = m.group(3).strip()
                    all_cols_seen.add(name + '_casa')
                    all_cols_seen.add(name + '_fora')

            all_rows.append(row)

            if writer is None:
                cols = ['mid'] + sorted(c for c in all_cols_seen if c != 'mid')
                writer = csv.DictWriter(f, fieldnames=cols, extrasaction='ignore')
                if not done:
                    writer.writeheader()

            writer.writerow(row)
            f.flush()
            print(f"OK ({len(seen_stats)} stats)")
        except Exception as e:
            print(f'ERRO: {e}')
            all_rows.append({'mid': mid})

        time.sleep(DELAY)

# ── Excel ──────────────────────────────────────────────────────────────────────
print('\nA gerar Excel...')
all_cols = ['mid'] + sorted(c for c in all_cols_seen if c != 'mid')

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Stats PPL 2025-26'

hdr_fill = PatternFill('solid', fgColor='004B87')   # azul PPL
hdr_font = Font(color='FFFFFF', bold=True)
for ci, col in enumerate(all_cols, 1):
    cell = ws.cell(row=1, column=ci, value=col)
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal='center')

for ri, row in enumerate(all_rows, 2):
    vals = []
    for col in all_cols:
        val = row.get(col, '')
        try:
            val = float(val.replace('%','').strip()) if val else None
        except Exception:
            pass
        vals.append(val)
    ws.append(vals)

for ci in range(1, len(all_cols)+1):
    ws.column_dimensions[get_column_letter(ci)].width = 18
ws.column_dimensions['A'].width = 12
ws.freeze_panes = 'B2'

wb.save(OUT_XLSX)
print(f'\n✓ CONCLUÍDO: {OUT_XLSX}')
print(f'  {len(all_rows)} jogos | {len(all_cols)} colunas')
