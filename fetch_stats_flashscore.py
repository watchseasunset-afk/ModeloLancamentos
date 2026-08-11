"""
Fetch ALL Flashscore stats para PL 2025/26 — 380 jogos
Grava em: data/stats_completas_PL_2025_26.xlsx
Corre: python fetch_stats_flashscore.py
Duração: ~8-10 min
"""
import time, re, csv, sys, subprocess
from pathlib import Path

# Instalar dependências
for pkg in ['requests', 'openpyxl']:
    try:
        __import__(pkg)
    except ImportError:
        subprocess.check_call([sys.executable, '-m', 'pip', 'install', pkg])

import requests, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── 380 Match IDs PL 2025/26 (Flashscore) ────────────────────────────────────
MATCH_IDS = """xQXUa3UG,buMwbsaT,UNC9hLMj,pWfdGOEc,Iqg4E2qA,zLXUefTr,roywfYce,40PohCS7,W6HXFFKE,CGPuEgkR,Uu6uknGb,rkXMW40U,Gxt6zm15,z1yE3oVi,MF2Xj8on,jeflmQpB,OM7eo4FN,WSUM1Pa4,r7lSurwo,AyZEYQVH,xfi4Ju8N,nclkvV1t,8Cxbx9Wh,UmvdMdte,Med3KzB7,xWhvMmnC,4IfWN9Ha,YLpmOIBr,2iZkrv3E,dY8uKRGO,8j1NPVnm,rTwrFuZR,n1pZG14F,jFEYCXop,zRIM7cR1,buItBBGd,6sw2IgOM,EmnRINZ2,tO58V0J8,xjaHTvlL,EBUD9Jdk,tQzy4f9T,OQsq6PYa,YaXS5GvH,h44gMyPc,bHZgKFgA,CpqK7xA4,0AH7bYWj,Khyh433C,6mtC9buh,jubPoemo,pbPN0UdK,U3M9iQJ6,xzTRt95l,GtGge8Ks,hx84D7tQ,42nywR4D,xAqy8o4m,djO1gnkf,QVFIk4lJ,dxqTvmz1,dG7hS9je,bLP5yET8,WYNDZibL,IVe1QmL7,pxlmliDE,jJ4pUVLr,AoUBFW6k,rDsJDAy2,jcHwYvNG,t4sdnVsR,U7CL8Og5,n5KMyNNi,fBXzEJVp,UTmdL1wB,0lduPqvn,8EamNN8b,bm4x5tgU,QZ5U62OH,0jR7cwU6,dtZvdoAa,0EuWc7um,UgXnfPvC,QLQeh39O,Ms9DApPh,4OXbaaaf,juf03THP,8GUFeHbJ,0SK4C6ft,vwIIJTp9,CWPWeQh2,Y1fIjVGM,QaZdqm9d,ELFi5UVF,jHyOcnOk,WAZj1LUs,IZVlo9vp,t8Hq7j13,KYph380S,CE2gREmB,lfdXJWuo,CSsSuE24,vqBFW9Pj,YsWzvhXG,42vKszYi,KbUrxW1T,Wxb1fDHc,hOA1PhIN,S4iAhinA,UBn9TqDs,rZAxVdln,bTewZ5kK,bivQP1C6,IgfRNusJ,QwtIRNsf,4laoTzJb,G8MZEpbl,CzGrDOS0,IXIjB2cD,rXDVZ0h4,lKNJm8ak,W2uAbPD7,4nnmgUyS,AP8E9kM8,2Nmg27Er,Ami10oqe,Kxy4K6zR,dpRRonU1,rw7M7TjL,OAseMS5E,lbnqyVFq,lrdmg9KU,xKSwcCx3,K24WdiL5,GUcueVkI,UyEecLWo,MqW2ea1b,dMruej7F,pQuhZ8pd,raKBgwWA,M7x6tWzP,CWHF0ZLt,YBkFXzhf,hSpkpEy0,Quqcrh6C,bTitnz7m,nkONThiJ,vPKFVEM6,Sp1AidqQ,Yo6ObDjg,neRXKc8l,YqgQO69M,4YS4HO1d,hf9aw6OF,tEAzuJaK,IHUdJpXq,KtBiuSg3,hhUuJyw1,E7IlHF7D,0bi7ZdNs,MBSHcUAN,QqZ8ajtB,QJkIQSvA,YVDiA7fT,IsEmh8Yp,dQISFBQi,4jMzEke4,xb9N4oIj,Uid9SlAc,Q5AqCTPG,SC5rmMjl,AZ3Bject,d4sh3Ytn,Ovaio0L0,W42aqvjD,r1HIlZRh,vTARnDd5,Ct9ZpiRH,882rqVeU,IJV01CBb,lMp9YMye,byvxt5qL,QNuUsRE8,AyZHW067,xrR3f4bE,tds1zq7r,0M7yxyRk,8nBqzFe2,WtKChrTQ,GlW4sc3S,EqdmT6rI,SEKlou43,844FWukA,UJhOUJJM,Kp3vVSD5,vgvfPcsp,pWIKAbSj,0v17Y1Kc,zLX2NyCd,OtEcqJYF,Qm9dKBam,j5BNYAEt,f5IQPXyQ,CpqN5Cpf,8CxW3jF6,E9uv2UqJ,MF25IkUa,CKaWWlqg,0tWDE7TO,fT0DGTbC,QPr9bewR,OtdbtdHr,hCp10H8E,pAPugZ0l,Kn4rDfhK,pdp7vzne,hMmFxEG7,hzTmiDV0,vXpF7YFs,EkHdki0D,n1p7rqmT,0pMPV53N,ILhjnRm4,6oRwwNAj,buEP0Q2c,IsUQb4X9,4ISGXRYA,dG7eQqRp,COkrlmJi,vysap5IG,WjF9Zm4b,l8cJOWkt,z1s3Mt9C,dY0RMAKh,GbR1O2v0,rRpBKKwP,jXgzLlk5,OphrJSJH,AyR0yAZo,roTgQO9m,0OUQMk3e,lWVnUXAL,8GCJOBYr,MeC5HfRF,EgipACJk,0Akh8jl2,lpBLrXPE,MwJZKTX7,ULFDFYeS,GIXvWgt9,Imk5ldcT,Eehlhvb4,8hnW2eCM,QTIdJGd3,EJzyqzLj,YZVAeD4d,vDfF6wDc,MLedjISG,nwbN4HsA,vyw2cZZq,Y1fLl5iC,xnqNuA8s,4M24hmhm,S2CgWdyo,AasVwlwf,6RzvxS76,fmHYc9UP,0ldDjRMa,ht4uf0Ti,Eivnz6xJ,fD0PWWVQ,8dxrKgnL,xCSDEX9r,6iKUAk97,xY3lRAgl,bRGMCBve,pb7dPlO0,Ma2HYh1E,YJmioWpK,6De5NShD,O82v0Yfk,WhxaK2t4,CUrZLFH8,OSYjMOBi,KzdmbCP1,EgB7VpZb,MJEFTO3B,GI7OR2YN,ClgfMtdo,h8Q7ItBG,EXZWI7kg,t6dj6U5m,EJfb48z0,Aol72n5C,hvNbX64n,2g6E0QZO,rDRmFPlI,67WOKTKt,2FskD3JU,AaTuHoK5,MLFJBjrQ,KUKzCax9,GOSXmmjJ,6kwlldp2,6Rf1c0NF,Eib9evhS,bPZujIFk,jHs8giMs,f9vPk9L6,nqyHiVjf,bkJWOuHj,6ivhWNOG,Kf4RpNwA,8KxFyJ0d,0EP0aMh3,Ec1Jnq9c,KztpYqf4,Chq6wuWq,tAhzq18M,2JX0U1gT,js9fNSun,CUIIHO9N,CWkRLF4s,bLiyz6Pi,WEakC2Xp,4n4zIX36,CC7rGBYI,b7C2L6Ab,vHKAJpuB,pplZJgZg,OUNtGjHL,fDECAUvq,4QSa5fK0,6VpDEbcE,bTPi7Gkl,GS2L889e,lfGS6nv8,QuH83YlD,hn9fRysK,vqqLCxSQ,tAlwdVPk,231BsWAd,Uk33qhtp,S2UXHCn9,zkoV10dB,tn6j5Z23,lC9b3DXF,ARc71i2S,z9snf9f2,SOsN3MRb,vLfbv6bm,KvGXh5Dt,CYh6xpTa,2N7lkNC5,GQkNY1SO,6eEuiqrg,zy6dm1sI,W2a5ouCU,69gE5rcn,80nFzOcC,fcMuMOrJ,QHUOP6qf,OYDhXliK,K4aORSEs,zFCHVyGF,vovkDTal,00ScB7U0,8QV59oaD,YsYWNpE6,M3ERqSzQ,ngWM9fgd,jRxEBGOq,thirHx1M,0YV82wnT,Kx3ZIbW9,KON04aHG,d4HvcH9j,CCaRKK1c,CbKh6Ln4,jqWq8sIi""".replace('\n','').split(',')

NINJA_URL = 'https://global.flashscore.ninja/20/x/feed/df_st_1_{mid}'
NINJA_HDR = {
    'X-Fsign': 'SW9D1eZo',
    'Referer': 'https://www.flashscore.com/',
    'Accept': '*/*',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
}
DELAY   = 1.5
DATA_DIR = Path(__file__).parent / 'data'
RAW_CSV  = DATA_DIR / 'stats_raw.csv'
OUT_XLSX = DATA_DIR / 'stats_completas_PL_2025_26.xlsx'
DATA_DIR.mkdir(exist_ok=True)

STAT_RE = re.compile(r'SG÷([^¬]+)¬SH÷([^¬]*)¬SI÷([^¬~]*)')

session = requests.Session()
session.headers.update(NINJA_HDR)

# ── Carregar progresso ─────────────────────────────────────────────────────────
done = set()
all_rows = []
if RAW_CSV.exists():
    with open(RAW_CSV, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            done.add(row['mid'])
            all_rows.append(dict(row))
    print(f'[resume] {len(done)} já recolhidos')

all_cols_seen = set()
for r in all_rows:
    all_cols_seen.update(r.keys())

# ── Scraping ───────────────────────────────────────────────────────────────────
todo = [m for m in MATCH_IDS if m not in done]
print(f'{len(todo)} jogos por recolher...\n')

with open(RAW_CSV, 'a', newline='', encoding='utf-8-sig') as f:
    writer = None

    for i, mid in enumerate(todo, 1):
        print(f'[{i}/{len(todo)}] {mid}', end=' ')
        try:
            r = session.get(NINJA_URL.format(mid=mid), timeout=15)
            r.raise_for_status()
            text = r.content.decode('utf-8', errors='replace')

            row = {'mid': mid}
            seen_stats = set()
            for m in STAT_RE.finditer(text):
                name = m.group(1).strip()
                if name not in seen_stats:  # primeira ocorrência = total jogo
                    seen_stats.add(name)
                    row[name + '_casa'] = m.group(2).strip()
                    row[name + '_fora'] = m.group(3).strip()
                    all_cols_seen.add(name + '_casa')
                    all_cols_seen.add(name + '_fora')

            all_rows.append(row)

            if writer is None:
                cols = ['mid'] + sorted(c for c in all_cols_seen if c != 'mid')
                writer = csv.DictWriter(f, fieldnames=cols, extrasaction='ignore')
                if not done:
                    writer.writeheader()

            writer.writerow(row)
            f.flush()
            print(f"OK ({len(seen_stats)} stats)")
        except Exception as e:
            print(f'ERRO: {e}')
            all_rows.append({'mid': mid})

        time.sleep(DELAY)

# ── Excel ──────────────────────────────────────────────────────────────────────
print('\nA gerar Excel...')
all_cols = ['mid'] + sorted(c for c in all_cols_seen if c != 'mid')

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Stats PL 2025-26'

hdr_fill = PatternFill('solid', fgColor='1F3864')
hdr_font = Font(color='FFFFFF', bold=True)
for ci, col in enumerate(all_cols, 1):
    cell = ws.cell(row=1, column=ci, value=col)
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal='center')

for ri, row in enumerate(all_rows, 2):
    vals = []
    for col in all_cols:
        val = row.get(col, '')
        try:
            val = float(val.replace('%','').strip()) if val else None
        except Exception:
            pass
        vals.append(val)
    ws.append(vals)

for ci in range(1, len(all_cols)+1):
    ws.column_dimensions[get_column_letter(ci)].width = 18
ws.column_dimensions['A'].width = 12
ws.freeze_panes = 'B2'

wb.save(OUT_XLSX)
print(f'\n✓ CONCLUÍDO: {OUT_XLSX}')
print(f'  {len(all_rows)} jogos | {len(all_cols)} colunas')
