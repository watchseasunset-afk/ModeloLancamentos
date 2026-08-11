"""
Calcula médias por equipa para PL, PPL e Brasil (Lançamentos + todas as stats).
Cria: data/medias_equipas_MULTI_LIGA.xlsx  (5 sheets por liga + sheet combinada)
Corre: python calc_medias_multi_liga.py
"""
import sys, subprocess, re
for pkg in ['pandas', 'openpyxl']:
    try: __import__(pkg)
    except ImportError: subprocess.check_call([sys.executable,'-m','pip','install',pkg,'--break-system-packages'])

import pandas as pd, numpy as np, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

DATA_DIR = Path(__file__).parent / 'data'
OUT_FILE = DATA_DIR / 'medias_equipas_MULTI_LIGA.xlsx'

LIGAS = [
    {'nome': 'PL',    'stats': 'stats_completas_PL_2025_26.xlsx',  'teams': Path(__file__).parent/'teams_map.csv',        'cor': '1F3864'},
    {'nome': 'PPL',   'stats': 'stats_completas_PPL_2025_26.xlsx', 'teams': Path(__file__).parent/'teams_map_ppl.csv',    'cor': '004B87'},
    {'nome': 'BRA',   'stats': 'stats_completas_BRA_2025.xlsx',    'teams': Path(__file__).parent/'teams_map_brazil.csv', 'cor': '009C3B'},
]

def parse_stat(val):
    if pd.isna(val): return np.nan
    s = str(val).strip()
    m = re.search(r'\((\d+)/(\d+)\)', s)
    if m: return float(m.group(2))
    try: return float(s.replace('%','').strip())
    except: return np.nan

def calc_medias_liga(liga):
    stats = pd.read_excel(DATA_DIR / liga['stats'])
    teams = pd.read_csv(liga['teams'])
    df = stats.merge(teams, on='mid', how='left')

    stat_cols = [c for c in stats.columns if c != 'mid']
    for c in stat_cols:
        df[c] = df[c].apply(parse_stat)

    stat_names = sorted(set(c.replace('_casa','').replace('_fora','') for c in stat_cols))
    teams_list = sorted(df['home'].dropna().unique())

    records_home, records_away, records_overall = [], [], []
    for team in teams_list:
        mh = df['home'] == team
        ma = df['away'] == team
        rh = {'Liga': liga['nome'], 'Equipa': team, 'Jogos_Casa': int(mh.sum())}
        ra = {'Liga': liga['nome'], 'Equipa': team, 'Jogos_Fora': int(ma.sum())}
        ro = {'Liga': liga['nome'], 'Equipa': team, 'Jogos_Total': int((mh|ma).sum())}

        for stat in stat_names:
            cc, cf = stat+'_casa', stat+'_fora'
            if cc in df.columns: rh[f'{stat}_media'] = df.loc[mh, cc].mean()
            if cf in df.columns: ra[f'{stat}_media'] = df.loc[ma, cf].mean()
            vals_h = df.loc[mh, cc] if cc in df.columns else pd.Series(dtype=float)
            vals_a = df.loc[ma, cf] if cf in df.columns else pd.Series(dtype=float)
            ro[f'{stat}_media'] = pd.concat([vals_h, vals_a]).mean()

        records_home.append(rh)
        records_away.append(ra)
        records_overall.append(ro)

    return (pd.DataFrame(records_home),
            pd.DataFrame(records_away),
            pd.DataFrame(records_overall),
            teams_list)

def style_sheet(ws, df, cor_hex):
    fill = PatternFill('solid', fgColor=cor_hex)
    font = Font(color='FFFFFF', bold=True, size=9)
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = fill; cell.font = font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    for ri, row in df.iterrows():
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri+2, column=ci)
            cell.value = round(float(val), 2) if isinstance(val, float) and not pd.isna(val) else val
            if (ri+2) % 2 == 0:
                cell.fill = PatternFill('solid', fgColor='EEF2F7')
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 22
    for ci in range(3, len(df.columns)+1):
        ws.column_dimensions[get_column_letter(ci)].width = 15
    ws.freeze_panes = 'C2'
    ws.row_dimensions[1].height = 28

print('A calcular médias por liga...')
wb = openpyxl.Workbook()
wb.remove(wb.active)

all_overall = []

for liga in LIGAS:
    print(f'  {liga["nome"]}...', end=' ')
    try:
        df_h, df_a, df_o, teams = calc_medias_liga(liga)
        all_overall.append(df_o)

        # Sheet Geral
        ws = wb.create_sheet(f'{liga["nome"]} - Geral')
        df_s = df_o.sort_values('Equipa').round(2)
        style_sheet(ws, df_s, liga['cor'])

        # Sheet Casa
        ws2 = wb.create_sheet(f'{liga["nome"]} - Casa')
        style_sheet(ws2, df_h.sort_values('Equipa').round(2), liga['cor'])

        # Sheet Fora
        ws3 = wb.create_sheet(f'{liga["nome"]} - Fora')
        style_sheet(ws3, df_a.sort_values('Equipa').round(2), liga['cor'])

        print(f'{len(teams)} equipas OK')
    except Exception as e:
        print(f'ERRO: {e}')

# Sheet combinada — só Lançamentos + Posse para comparação inter-liga
print('  A criar sheet comparativa...')
try:
    dfc = pd.concat(all_overall, ignore_index=True)
    cols_keep = ['Liga','Equipa','Jogos_Total']
    for stat in ['Lançamentos_media','Posse de bola_media','Passes longos_media',
                 'Alívios_media','Cruzamentos_media','Faltas_media']:
        if stat in dfc.columns:
            cols_keep.append(stat)
    dfc_s = dfc[[c for c in cols_keep if c in dfc.columns]].sort_values(['Liga','Equipa'])
    wsc = wb.create_sheet('Comparativa Inter-Liga', 0)
    style_sheet(wsc, dfc_s.reset_index(drop=True), '2E4057')
except Exception as e:
    print(f'  Comparativa ERRO: {e}')

wb.save(OUT_FILE)
print(f'\n✓ CONCLUÍDO: {OUT_FILE}')
print(f'  Sheets: {[s.title for s in wb.sheetnames if hasattr(s,"title")]}')
for ws in wb.worksheets:
    print(f'  - {ws.title}')
